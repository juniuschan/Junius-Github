import openpyxl


sys_file = 'ch167.xlsx'  # load_data()
deli_note = 'sales_record.xlsx'
certificate = 'certificate.xlsx'


# B E D S T U L N F G !! H
def load_data(xlsxName):
    workbook = openpyxl.load_workbook(xlsxName)
    worksheet = workbook['Sheet1']    # sheet name should be 'Sheet1'
    data_result = []
    for row in range(2, worksheet.max_row + 1):
        row_data = []
        row_data.append(worksheet['B' + str(row)].value)
        row_data.append(worksheet['E' + str(row)].value)
        row_data.append(worksheet['D' + str(row)].value)
        row_data.append(worksheet['S' + str(row)].value)
        row_data.append(worksheet['T' + str(row)].value)
        row_data.append(worksheet['U' + str(row)].value)
        row_data.append(worksheet['L' + str(row)].value)
        row_data.append(worksheet['N' + str(row)].value)
        row_data.append(worksheet['F' + str(row)].value)
        row_data.append(worksheet['G' + str(row)].value)
        row_data.append(worksheet['H' + str(row)].value)
        data_result.append(row_data)
        del row_data
    print("DONE !")
    return data_result


data_dict = load_data(sys_file)
print("data_dict max rows: ", len(data_dict))
# for i in range(len(data_dict)):
#     print(data_dict[i])


workbook = openpyxl.load_workbook(deli_note)
worksheet = workbook['Sheet1']

#for row in (4, 5, 6, 7, 15, 16, 17, 18, 27, 28, 29, 30, 38, 39, 40, 41):
#Edit your starting row here:
""" EDIT YOUR STARTING ROW HERE """
r_1 = 4
#r_2 = 15
#r_3 = 27
#r_4 = 38

#name, model, unit, quan, price, total, lot, expr, comp, logNo, store = data_dict[i]
for i in range(4):
    """ Form One """
    worksheet.cell(row=r_1 + i, column=2).value = data_dict[0 + i][0]
    worksheet.cell(row=r_1 + i, column=3).value = data_dict[0 + i][1]
    worksheet.cell(row=r_1 + i, column=4).value = data_dict[0 + i][2]
    worksheet.cell(row=r_1 + i, column=5).value = data_dict[0 + i][3]
    worksheet.cell(row=r_1 + i, column=6).value = data_dict[0 + i][4]
    worksheet.cell(row=r_1 + i, column=7).value = data_dict[0 + i][5]
    worksheet.cell(row=r_1 + i, column=8).value = data_dict[0 + i][6]
    worksheet.cell(row=r_1 + i, column=10).value = data_dict[0 + i][7]
    worksheet.cell(row=r_1 + i, column=11).value = data_dict[0 + i][8]
    worksheet.cell(row=r_1 + i, column=12).value = data_dict[0 + i][9]
    worksheet.cell(row=r_1 + i, column=14).value = data_dict[0 + i][10]
    # """ Form Two """
    # worksheet.cell(row=r_2 + i, column=2).value = data_dict[4 + i][0]
    # worksheet.cell(row=r_2 + i, column=3).value = data_dict[4 + i][1]
    # worksheet.cell(row=r_2 + i, column=4).value = data_dict[4 + i][2]
    # worksheet.cell(row=r_2 + i, column=5).value = data_dict[4 + i][3]
    # worksheet.cell(row=r_2 + i, column=6).value = data_dict[4 + i][4]
    # worksheet.cell(row=r_2 + i, column=7).value = data_dict[4 + i][5]
    # worksheet.cell(row=r_2 + i, column=8).value = data_dict[4 + i][6]
    # worksheet.cell(row=r_2 + i, column=10).value = data_dict[4 + i][7]
    # worksheet.cell(row=r_2 + i, column=11).value = data_dict[4 + i][8]
    # worksheet.cell(row=r_2 + i, column=12).value = data_dict[4 + i][9]
    # worksheet.cell(row=r_2 + i, column=14).value = data_dict[4 + i][10]
    # """ Form Three """
    # worksheet.cell(row=r_3 + i, column=2).value = data_dict[8 + i][0]
    # worksheet.cell(row=r_3 + i, column=3).value = data_dict[8 + i][1]
    # worksheet.cell(row=r_3 + i, column=4).value = data_dict[8 + i][2]
    # worksheet.cell(row=r_3 + i, column=5).value = data_dict[8 + i][3]
    # worksheet.cell(row=r_3 + i, column=6).value = data_dict[8 + i][4]
    # worksheet.cell(row=r_3 + i, column=7).value = data_dict[8 + i][5]
    # worksheet.cell(row=r_3 + i, column=8).value = data_dict[8 + i][6]
    # worksheet.cell(row=r_3 + i, column=10).value = data_dict[8 + i][7]
    # worksheet.cell(row=r_3 + i, column=11).value = data_dict[8 + i][8]
    # worksheet.cell(row=r_3 + i, column=12).value = data_dict[8 + i][9]
    # worksheet.cell(row=r_3 + i, column=14).value = data_dict[8 + i][10]
    # """ Form Four """
    # worksheet.cell(row=r_4 + i, column=2).value = data_dict[12 + i][0]
    # worksheet.cell(row=r_4 + i, column=3).value = data_dict[12 + i][1]
    # worksheet.cell(row=r_4 + i, column=4).value = data_dict[12 + i][2]
    # worksheet.cell(row=r_4 + i, column=5).value = data_dict[12 + i][3]
    # worksheet.cell(row=r_4 + i, column=6).value = data_dict[12 + i][4]
    # worksheet.cell(row=r_4 + i, column=7).value = data_dict[12 + i][5]
    # worksheet.cell(row=r_4 + i, column=8).value = data_dict[12 + i][6]
    # worksheet.cell(row=r_4 + i, column=10).value = data_dict[12 + i][7]
    # worksheet.cell(row=r_4 + i, column=11).value = data_dict[12 + i][8]
    # worksheet.cell(row=r_4 + i, column=12).value = data_dict[12 + i][9]
    # worksheet.cell(row=r_4 + i, column=14).value = data_dict[12 + i][10]


#
#
wb_cert = openpyxl.load_workbook(certificate)
ws_cert = wb_cert['Sheet1']
# the data information in certificate.xlsx
cert_data = {}

for row in range(2, ws_cert.max_row + 1):
    name = ws_cert['H' + str(row)].value
    exp_ = ws_cert['AF' + str(row)].value
    cert_data.setdefault(name, exp_)

# workbook = openpyxl.load_workbook(deli_note)
# worksheet = workbook['Sheet1']
for nRow in range(2, worksheet.max_row + 1):
    name = worksheet.cell(nRow, column = 2).value
    if name in cert_data:
        worksheet.cell(nRow, column = 13).value = cert_data[name]

workbook.save(deli_note)
# #2 3 4 5 6 7 8 10 11 12 14
# test = load_data()
# print(test[3])
# for i in range(len(test)):
#     print(test[i])
