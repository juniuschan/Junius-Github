import openpyxl


"""
a demo.xlsx and sales_rec.xlsx are needed specifically
"""
sys_rec = 'demo.xlsx'
manu_rec = 'sales_rec.xlsx'
CERTIFICATE = 'certificate.xlsx' # set to be a constant


# worksheet name should be 'Sheet1'
def get_worksheet(xlsxName):
    workbook = openpyxl.load_workbook(xlsxName)
    return workbook['Sheet1']

# B E D S T U L N F G !! H
def load_data(worksheet):
    for row in range(2, worksheet.max_row + 1):
        row_data = []
        row_data.append(worksheet['B' + str(row)].value)
        row_data.append(worksheet['E' + str(row)].value)
        row_data.append(worksheet['D' + str(row)].value)
        row_data.append(worksheet['S' + str(row)].value)
        row_data.append(worksheet['T' + str(row)].value)
        row_data.append(worksheet['U' + str(row)].value)
        row_data.append(worksheet['L' + str(row)].value)
        row_data.append(worksheet['N' + str(row)].value)
        row_data.append(worksheet['F' + str(row)].value)
        row_data.append(worksheet['G' + str(row)].value)
        row_data.append(worksheet['H' + str(row)].value)
        yield row_data


def write2xlsx():
    num = 1
    workbook = openpyxl.load_workbook(manu_rec)
    worksheet = workbook['Sheet1']
    for data in load_data(get_worksheet(sys_rec)):
        worksheet.cell(row=num, column=2).value = data[0]
        worksheet.cell(row=num, column=3).value = data[1]
        worksheet.cell(row=num, column=4).value = data[2]
        worksheet.cell(row=num, column=5).value = data[3]
        worksheet.cell(row=num, column=6).value = data[4]
        worksheet.cell(row=num, column=7).value = data[5]
        worksheet.cell(row=num, column=8).value = data[6]
        worksheet.cell(row=num, column=10).value = data[7]
        worksheet.cell(row=num, column=11).value = data[8]
        worksheet.cell(row=num, column=12).value = data[9]
        worksheet.cell(row=num, column=14).value = data[10]
        num += 1
    workbook.save(manu_rec)


def update_xlsx():
    worksheet = get_worksheet(CERTIFICATE)
    cert_data = {}
    for row in range(2, worksheet.max_row + 1):
        name = worksheet['H' + str(row)].value
        exp_date = worksheet['AF' + str(row)].value
        cert_data.setdefault(name, exp_date)
    workbook = openpyxl.load_workbook(manu_rec)
    worksheet = workbook['Sheet1']
    for row in range(1, worksheet.max_row + 1):
        name = worksheet.cell(row, column = 2).value
        if name in cert_data:
            worksheet.cell(row, column = 13).value = cert_data[name]
    workbook.save(manu_rec)


def main():
    write2xlsx()
    update_xlsx()
    print("FINISH !")

if __name__ == "__main__":
    main()

